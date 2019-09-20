import os
from datetime import datetime, timedelta

import openpyxl
import xlwt
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.shortcuts import reverse
from django.test import Client, TestCase
from django.test.utils import override_settings
from haystack import connections

from common.models import User
from marketing.models import (Campaign, CampaignLinkClick, CampaignLog,
                              CampaignOpen, Contact, ContactEmailCampaign,
                              ContactList, EmailTemplate, FailedContact, Link,
                              Tag)
from marketing.views import *

TEST_INDEX = {
    'default': {
        'ENGINE': 'haystack.backends.elasticsearch_backend.ElasticsearchSearchEngine',
        'URL': 'http://127.0.0.1:9200/',
        'TIMEOUT': 60 * 10,
        'INDEX_NAME': 'test_index',
    },
}

class TestMarketingModel(object):
    def setUp(self):
        self.client = Client()

        self.user = User.objects.create(
            username='john', email='john@example.com', role="ADMIN")
        self.user.set_password('password')
        self.user.save()

        self.user1 = User.objects.create(
            username='janeMarketing', email='janeMarketing@example.com', role="USER", has_marketing_access=True)
        self.user1.set_password('password')
        self.user1.save()

        self.client.login(username='john@example.com', password='password')

        self.tag_marketing = Tag.objects.create(name='tag marketing', created_by=self.user)
        self.tag_marketing_1 = Tag.objects.create(name='tag marketing _1')

        self.contact_list = ContactList.objects.create(
            name='contacts_list by admin', created_by=self.user)
        self.contact_list.tags.add(self.tag_marketing.id)

        self.contact_list_user = ContactList.objects.create(
            name='contacts_list by user', created_by=self.user1)

        self.contact_list_user_1 = ContactList.objects.create(
            name='contacts_list by user 1', created_by=self.user1)

        self.contact = Contact.objects.create(
            name='john doe', email='johnDoe@email.com', created_by=self.user)
        self.contact_1 = Contact.objects.create(
            name='jane doe', email='janeDoe@email.com', is_bounced=True)
        self.contact_1.contact_list.add(
            self.contact_list.id, self.contact_list_user.id)
        self.contact_2 = Contact.objects.create(
            name='joe doe', email='joeDoe@email.com')
        self.contact_3 = Contact.objects.create(
            name='jill doe', email='jillDoe@email.com')
        self.contact_4 = Contact.objects.create(
            name='jill Doe', email='jilldoe@email.com')
        self.contact_4.contact_list.add(self.contact_list.id)
        self.contact_5 = Contact.objects.create(
            name='jack Doe', email='jack doe@email.com',
            is_bounced=True)
        self.contact_5.contact_list.add(self.contact_list.id)
        self.contact_6 = Contact.objects.create(
            name='contact 6', email='contact6@email.com')
        self.contact_6.contact_list.add(self.contact_list.id)
        self.failed_contact = FailedContact.objects.create(
            name='jill Doe', email='jilldoe@email.com')
        self.failed_contact.contact_list.add(self.contact_list.id)

        self.email_template = EmailTemplate.objects.create(
            created_by=self.user, title='email title', subject='email subject',
            html='email html body',
        )

        self.campaign = Campaign.objects.create(
            title='campaign object',
            created_by=self.user,
            email_template=self.email_template,
            reply_to_email='django@crm.io',
            subject='subject of the campaign',
            html='html content',
            from_email='from@email.com',
            from_name='from name',
        )

        self.campaign_scheduled_later = Campaign.objects.create(
            title='campaign object',
            created_by=self.user,
            email_template=self.email_template,
            reply_to_email='django@crm.io',
            subject='subject of the campaign',
            from_email='from@email.com',
            from_name='from name',
            timezone='Asia/Kolkata',
            status='Scheduled',
            schedule_date_time=(datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d %H:%M'),
            html='<p><a href="http://example.com" target="_blank">links</a></p>'
        )

        self.campaign.contact_lists.add(self.contact_list.id, self.contact_list_user.id)
        self.campaign.tags.add(self.tag_marketing.id)

        self.link = Link.objects.create(
            campaign=self.campaign,
            original='https://example.com',
        )

        self.campaign_link_click = CampaignLinkClick.objects.create(
            campaign=self.campaign,
            link=self.link,
            contact=self.contact,
            ip_address='127.0.0.1'
        )

        self.contact_email_campaign = ContactEmailCampaign.objects.create(
            name='conta@admin', email='contactEmail@admin.com'
        )

        self.contact_email_campaign_1 = ContactEmailCampaign.objects.create(
            name='another@admin', email='contactEmailAnother@admin.com'
        )

        self.campaign_log = CampaignLog.objects.create(
            campaign=self.campaign,
            contact=self.contact_1
        )
        # connections.reload('default')

    # def tearDown(self):
    #     call_command('clear_index', interactive=False, verbosity=0)


@override_settings(HAYSTACK_CONNECTIONS=TEST_INDEX )
class TestTemplates(TestMarketingModel, TestCase):

    def setUp(self):
        super(TestTemplates, self).setUp()
        connections.reload('default')

    def test_templates(self):
        url1 = reverse('marketing:dashboard')
        url2 = reverse('marketing:contact_lists')
        url4 = reverse('marketing:contacts_list')
        url5 = reverse('marketing:contact_list_new')
        url8 = reverse('marketing:email_template_list')
        url9 = reverse('marketing:email_template_new')
        url12 = reverse('marketing:campaign_list')
        url13 = reverse('marketing:campaign_new')
        resp1 = self.client.get(url1)
        resp2 = self.client.get(url2)
        # resp4 = self.client.get(url4)
        resp5 = self.client.get(url5)
        resp8 = self.client.get(url8)
        resp9 = self.client.get(url9)
        # resp12 = self.client.get(url12)
        resp13 = self.client.get(url13)
        self.assertEqual(resp1.status_code, 200)
        self.assertEqual(resp2.status_code, 200)
        # self.assertEqual(resp4.status_code, 200)
        self.assertEqual(resp5.status_code, 200)
        self.assertEqual(resp8.status_code, 200)
        self.assertEqual(resp9.status_code, 200)
        # self.assertEqual(resp12.status_code, 200)
        self.assertEqual(resp13.status_code, 200)


class TestCreateContacts(TestMarketingModel, TestCase):

    def test_contact_list_new(self):
        data = ['company name,email,first name,last name,city,state\n',
                'mp,admin@mp,Admin,MP,Hyderabad,Telangana\n',
                'mp,hrmp.com,HR,MP,Hyderabad,Telangana\n',
                'mp,contactus@mp.com,,MP,Hyderabad,Telangana\n',
                'mp,test@mp.com,Test,MP,Hyderabad,Telangana\n',
                'mp,hello@mp.com,Hello,MP,Hyderabad,Telangana\n']

        with open('marketing/test_file.csv', 'w') as fp:
            fp.writelines(data)

        with open('marketing/test_file.csv') as fp:
            response = self.client.post(reverse('marketing:contact_list_new'),
                                        {'name': 'sample_test', 'attachment': fp})
            self.assertEqual(response.status_code, 200)

        os.remove('marketing/test_file.csv')


class TestViewContactList(TestMarketingModel, TestCase):

    # def test_view_contact_list(self):
    #     ## ContactList create object first !!!
    #     response = self.client.get('/m/cl/list/1/detail/')
    #     self.assertEqual(response.status_code, 200)
    #     self.assertTemplateUsed(response, 'marketing/lists/detail.html')

    def test_contact_list_pagination(self):
        response = self.client.get(
            reverse('marketing:contact_lists') + '?page=1')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:contact_lists') + '?page=asdf')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:contact_lists') + '?page=')
        self.assertEqual(response.status_code, 200)


class TestEmailTemplateList(TestMarketingModel, TestCase):

    def test_email_template_list_pagination(self):
        response = self.client.get(
            reverse('marketing:email_template_list') + '?page=1')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:email_template_list') + '?page=asdf')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:email_template_list') + '?page=')
        self.assertEqual(response.status_code, 200)


class TestDasboardView(TestMarketingModel, TestCase):

    def test_marketing_dashboard(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:dashboard'))
        self.assertEqual(response.status_code, 200)

        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:dashboard'))
        self.assertEqual(response.status_code, 200)


@override_settings(HAYSTACK_CONNECTIONS=TEST_INDEX )
class TestContactListsListPage(TestMarketingModel, TestCase):

    def setUp(self):
        super(TestContactListsListPage, self).setUp()
        connections.reload('default')

    def test_contact_lists_list_page(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_lists'))
        self.assertEqual(response.status_code, 200)

        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_lists') + '?tag={}'.format(self.tag_marketing.id))
        self.assertEqual(response.status_code, 200)

        data = {
            'contact_list_name': 'name of contact list',
            'created_by': self.user1.id,
            'tag': self.tag_marketing.id,
        }
        response = self.client.post(
            reverse('marketing:contact_lists'), data)
        self.assertEqual(response.status_code, 200)


@override_settings(HAYSTACK_CONNECTIONS=TEST_INDEX )
class TestContactsListPage(TestMarketingModel, TestCase):

    def setUp(self):
        super(TestContactsListPage, self).setUp()
        connections.reload('default')
        # connections = ConnectionHandler(TEST_INDEX )

    def test_contacts_list_page(self):
        connections.reload('default')
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contacts_list'))
        self.assertEqual(response.status_code, 200)

        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contacts_list'))
        self.assertEqual(response.status_code, 200)

        data = {
            'contact_name': 'name of contact',
            'created_by': self.user1.id,
            'contact_email': 'contact@email.com',
            'contact_list': self.contact_list.id,
        }
        response = self.client.post(
            reverse('marketing:contacts_list'), data)
        self.assertEqual(response.status_code, 200)


class TestContactsCSVFileUploadView(TestMarketingModel, TestCase):

    def test_contacts_csv_file_upload_view(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_new'))
        self.assertEqual(response.status_code, 200)

        file_content = [
            'company name,email,first name,last name,city,state\n',
            'company_name_1,user1@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_2,user2@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_3,user3@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_4,user4@email.com,first_name,last_name,Hyderabad,Telangana\n'
        ]
        file_content = bytes(''.join(file_content), 'utf-8')
        data = {'contacts_file': SimpleUploadedFile(
            'file name', file_content), 'name': 'contact file name', 'tags': 'tag1,tag2,tag marketing'}
        response = self.client.post(
            reverse('marketing:contact_list_new'), data)
        self.assertEqual(201, response.status_code)

        # invalid emails
        file_content = [
            'company name,email,first name,last name,city,state\n',
            'company_name_1,user1@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_2,useremail.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_3,user3@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_4,user5@email,first_name,last_name,Hyderabad,Telangana\n'
        ]
        file_content = bytes(''.join(file_content), 'utf-8')
        data = {'contacts_file': SimpleUploadedFile(
            'file name', file_content), 'name': 'contact file name', 'tags': 'tag1,tag2,tag marketing'}
        response = self.client.post(
            reverse('marketing:contact_list_new'), data)
        self.assertEqual(200, response.status_code)

        # missing email header
        file_content = [
            'company name,first name,last name,city,state\n',
            'company_name_1,user1@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_2,user2@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_3,user3@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_4,user4@email.com,first_name,last_name,Hyderabad,Telangana\n'
        ]
        file_content = bytes(''.join(file_content), 'utf-8')
        data = {'contacts_file': SimpleUploadedFile(
            'file name', file_content), 'name': 'contact file name', 'tags': 'tag1,tag2,tag marketing'}
        response = self.client.post(
            reverse('marketing:contact_list_new'), data)
        self.assertEqual(200, response.status_code)

        # missing all required headers
        file_content = [
            'company_name_1,user1@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_2,user2@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_3,user3@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_4,user4@email.com,first_name,last_name,Hyderabad,Telangana\n'
        ]
        file_content = bytes(''.join(file_content), 'utf-8')
        data = {'contacts_file': SimpleUploadedFile(
            'file name', file_content), 'name': 'contact file name', 'tags': 'tag1,tag2,tag marketing'}
        response = self.client.post(
            reverse('marketing:contact_list_new'), data)
        self.assertEqual(200, response.status_code)


class TestEditContactsCSVFileUploadView(TestMarketingModel, TestCase):

    def test_edit_contacts_csv_file_upload_view(self):
        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:edit_contact_list', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 200)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:edit_contact_list', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 200)

        file_content = [
            'company name,email,first name,last name,city,state\n',
            'company_name_1,user1@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_2,user2@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_3,user3@email.com,first_name,last_name,Hyderabad,Telangana\n',
            'company_name_4,user4@email.com,first_name,last_name,Hyderabad,Telangana\n'
        ]
        file_content = bytes(''.join(file_content), 'utf-8')
        data = {'contacts_file': SimpleUploadedFile(
            'file name', file_content), 'name': 'contacts_list by admin', 'tags': 'tag1,tag2,tag marketing'}
        response = self.client.post(
            reverse('marketing:edit_contact_list', args=(self.contact_list.id,)), data)
        self.assertEqual(200, response.status_code)

        data = {**data, 'name': ''}
        response = self.client.post(
            reverse('marketing:edit_contact_list', args=(self.contact_list.id,)), data)
        self.assertEqual(200, response.status_code)


class TestEditContact(TestMarketingModel, TestCase):

    def test_edit_contact(self):
        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:edit_contact', args=(self.contact.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:edit_contact', args=(self.contact.id,)))
        self.assertEqual(response.status_code, 200)

        data = {
            'name': 'John DOE',
            'email': self.contact.email
        }

        response = self.client.post(
            reverse('marketing:edit_contact', args=(self.contact.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name': 'John DOE',
            'email': self.contact.email,
            'from_url': self.contact_list.id,
        }

        response = self.client.post(
            reverse('marketing:edit_contact', args=(self.contact.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name': 'John DOE',
            'email': self.contact_4.email,
            'from_url': self.contact_list.id,
        }

        response = self.client.post(
            reverse('marketing:edit_contact', args=(self.contact.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name': 'John DOE',
            'email': ''
        }

        response = self.client.post(
            reverse('marketing:edit_contact', args=(self.contact.id,)) + '?from_url={}'.format(self.contact_list.id,), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name': 'JAne DOE',
            'email': 'janeDoe@mark.com',
            'from_url': self.contact_list.id,
        }

        response = self.client.post(
            reverse('marketing:edit_contact', args=(self.contact_1.id,)), data)
        self.assertEqual(response.status_code, 200)


class TestDeleteContact(TestMarketingModel, TestCase):

    def test_delete_contact(self):
        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_contact', args=(self.contact.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_contact', args=(self.contact_2.id,)))
        self.assertEqual(response.status_code, 302)

        response = self.client.get(
            reverse('marketing:delete_contact', args=(self.contact_3.id,)
                    ) + '?from_contact={}'.format(self.contact_list.id))
        self.assertEqual(response.status_code, 302)

        response = self.client.get(
            reverse('marketing:delete_contact', args=(self.contact_4.id,)
                    ) + '?from_contact={}'.format(self.contact_list_user.id))
        self.assertEqual(response.status_code, 302)


class TestContactListDetail(TestMarketingModel, TestCase):

    def test_contact_list_detail(self):
        self.client.login(
            username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_detail', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_detail', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 200)

        data = {
            'name':'contact name',
            'email':'contact@email.com',
            'company_name':'company name'
        }

        response = self.client.post(
            reverse('marketing:contact_list_detail', args=(self.contact_list.id,)), data)
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            reverse('marketing:contact_list_detail', args=(self.contact_list.id,)
            ) + '?bounced_contacts_page=0', data)
        self.assertEqual(response.status_code, 200)

        # failed contacts detail test
        # response = self.client.get(
        #     reverse('marketing:failed_contact_list_detail', args=(self.contact_list.id,)))
        # self.assertEqual(response.status_code, 200)

    def test_failed_contact_list_download_delete(self):
        response = self.client.get(
            reverse('marketing:failed_contact_list_download_delete', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:failed_contact_list_download_delete', args=(self.contact_list_user.id,)))
        self.assertEqual(response.status_code, 302)

    def test_email_template_list(self):
        response = self.client.get(
            reverse('marketing:email_template_list'))
        self.assertEqual(response.status_code, 200)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_detail', args=(self.contact_list.id,)))

        self.assertEqual(response.status_code, 200)
        data = {
            'template_name':'template name',
            'created_by':self.user.id,
        }

        response = self.client.post(
            reverse('marketing:email_template_list'), data)
        self.assertEqual(response.status_code, 200)

        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_list'))
        self.assertEqual(response.status_code, 200)

    def test_email_template_create(self):
        data = {
            'title': '',
            'subject': '',
            'html': '',
        }
        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:email_template_new'), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'title': 'title',
            'subject': 'subject',
            'html': 'html',
        }

        response = self.client.post(
            reverse('marketing:email_template_new'), data)
        self.assertEqual(response.status_code, 201)

    def test_email_edit(self):
        data = {
            'title': '',
            'subject': '',
            'html': '',
        }

        response = self.client.get(
            reverse('marketing:email_template_edit', args=(self.email_template.id,)), data)
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            reverse('marketing:email_template_edit', args=(self.email_template.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'title': 'title',
            'subject': 'subject',
            'html': 'html',
        }

        response = self.client.post(
            reverse('marketing:email_template_edit', args=(self.email_template.id,)), data)
        self.assertEqual(response.status_code, 201)

        self.client.login(username='janeMarketing@example.com', password='password')

        data = {
            'title': 'title',
            'subject': 'subject',
            'html': 'html',
        }

        response = self.client.post(
            reverse('marketing:email_template_edit', args=(self.email_template.id,)), data)
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_edit', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 200)


    def test_email_detail(self):
        self.client.logout()
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_detail', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_detail', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 200)


    def test_email_delete(self):
        self.client.logout()
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_delete', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:email_template_delete', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 302)

    def test_campaign_list(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_list') + '?tag={}'.format(self.tag_marketing.id))
        self.assertEqual(response.status_code, 200)

        data = {
            'campaign_name':'name of the campaign',
            'created_by':self.user.id,
        }

        response = self.client.post(
            reverse('marketing:campaign_list'), data)
        self.assertEqual(response.status_code, 200)

    def test_campaign_new(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_new'))
        self.assertEqual(response.status_code, 200)

        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_new'))
        self.assertEqual(response.status_code, 200)

        data = {
            'title': 'campaign title',
            'email_template': self.email_template.id,
            'contact_list':self.contact_list.id,
            'subject': 'campaign subject',
            'from_name': 'from name',
            'from_email': 'from@email.com',
            'reply_to_email': 'noreply@example.com',
            'tags': 'tag,marketing,tag marketing,new tag create',
            'timezone': 'Asia/Kolkata',
            'html': '<h1>html body</h1><p><br></p><p><strong>message body</strong></p><p><br></p><p><a href="http://example.com" target="_blank">links</a></p><p><br></p><p><br></p>'
        }

        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:campaign_new'), data)
        self.assertEqual(response.status_code, 201)

        data = {
            'title': 'campaign title',
            'email_template': self.email_template.id,
            'contact_list':self.contact_list.id,
            'subject': 'campaign subject',
            'from_name': 'from name',
            'from_email': 'from@email.com',
            'reply_to_email': 'noreply@example.com',
            'tags': 'tag,marketing,tag marketing,new tag create',
            'timezone': 'Asia/Kolkata',
            'schedule_later': 'true',
            'schedule_date_time': (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d %H:%M'),
            'html': '<h1>html body</h1><p><br></p><p><strong>message body</strong></p><p><br></p><p><a href="http://example.com" target="_blank">links</a></p><p><br></p><p><br></p>'
        }

        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:campaign_new'), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'title': 'campaign title new',
            'email_template': self.email_template.id,
            'contact_list':self.contact_list.id,
            'subject': 'campaign subject',
            'from_name': 'from name',
            'from_email': 'from@email.com',
            'reply_to_email': 'noreply@example.com',
            'tags': 'tag,marketing,tag marketing,new tag create',
            'timezone': 'Asia/Kolkata',
            'schedule_later': 'true',
            'schedule_date_time': (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d %H:%M'),
            'reply_to_crm':'true',
            'reply_to_email':'django@crm.com',
            'html': '<h1>html body</h1><p><br></p><p><strong>message body</strong></p><p><br></p><p><a href="http://example.com" target="_blank">links</a></p><p><br></p><p><br></p>'
        }

        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:campaign_new'), data)
        self.assertEqual(response.status_code, 201)

        data= { **data, 'title': 'wrong schedule date time',
            'schedule_date_time': (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d %H:%M'),
            'contact_list':self.contact_list_user_1.id,
            'html':'{{}} {{'
            }
        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:campaign_new'), data)
        self.assertEqual(response.status_code, 200)

        data= { **data, 'title': 'wrong schedule date time',
            'schedule_date_time': (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d %H:%M'),
            'contact_list':self.contact_list_user_1.id,
            'html':'}}'
            }
        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:campaign_new'), data)
        self.assertEqual(response.status_code, 200)

    def test_campaign_detail(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_details', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_details', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:campaign_details', args=(self.campaign.id,)
            ) + '?page=1')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:campaign_details', args=(self.campaign.id,)
            ) + '?page=None')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:campaign_details', args=(self.campaign.id,)
            ) + '?page=')
        self.assertEqual(response.status_code, 200)

    def test_campaign_delete(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_delete', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_delete', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 302)

    def test_demo_file_download(self):
        response = self.client.get(
            reverse('marketing:demo_file_download'))
        self.assertEqual(response.status_code, 200)

    def test_unsubscribe_from_campaign(self):
        response = self.client.get(
            reverse('marketing:unsubscribe_from_campaign', kwargs={'contact_id': self.contact.id,'campaign_id': self.campaign.id,}))
        self.assertEqual(response.status_code, 200)

    def test_contact_detail(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_detail', args=(self.contact.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_detail', args=(self.contact.id,)))
        self.assertEqual(response.status_code, 200)

    def test_download_contacts_for_campaign(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:download_contacts_for_campaign', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:download_contacts_for_campaign', args=(self.campaign.id,)
            ) + '?is_bounced=true')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:download_contacts_for_campaign', args=(self.campaign.id,)
            ) + '?is_unsubscribed=true')
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:download_contacts_for_campaign', args=(self.campaign.id,)
            ) + '?is_opened=true')
        self.assertEqual(response.status_code, 200)

    def test_create_campaign_from_template(self):
        response = self.client.get(
            reverse('marketing:create_campaign_from_template', args=(self.email_template.id,)))
        self.assertEqual(response.status_code, 302)

    def test_download_link_clicked(self):
        response = self.client.get(
            reverse('marketing:download_links_clicked', args=(self.campaign.id,)))
        self.assertEqual(response.status_code, 200)

    def test_delete_multiple_contacts(self):
        data = {
            'selected_list[]':[self.contact.id, self.contact_1.id],
            'from_contact':self.contact_list.id,
        }
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.post(
            reverse('marketing:delete_multiple_contacts'), data)
        self.assertEqual(response.status_code, 200)

        self.client.login(username='john@example.com', password='password')
        response = self.client.post(
            reverse('marketing:delete_multiple_contacts'), data)
        self.assertEqual(response.status_code, 200)

    def test_delete_all_contacts(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_all_contacts', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_all_contacts', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 302)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_all_contacts', args=(self.contact_list.id,)
            ) + '?bounced=true')
        self.assertEqual(response.status_code, 302)

    def test_download_failed_contacts(self):
        self.client.login(username='janeMarketing@example.com', password='password')
        response = self.client.get(
            reverse('marketing:download_failed_contacts', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 403)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:download_failed_contacts', args=(self.contact_list.id,)))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(
            reverse('marketing:download_failed_contacts', args=(self.contact_list_user_1.id,)))
        self.assertEqual(response.status_code, 200)

    def test_list_all_email_for_campaigns(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:list_all_emails_for_campaigns'))
        self.assertEqual(response.status_code, 200)

    def test_add_email_for_campaigns(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:add_email_for_campaigns'))
        self.assertEqual(response.status_code, 200)

        data = {}

        response = self.client.post(
            reverse('marketing:add_email_for_campaigns'), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name':'admin name',
            'email':'admin@email.com'
        }

        response = self.client.post(
            reverse('marketing:add_email_for_campaigns'), data)
        self.assertEqual(response.status_code, 200)

    def test_edit_email_for_campaigns(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:edit_email_for_campaigns', args=(self.contact_email_campaign.id,)))
        self.assertEqual(response.status_code, 200)

        data = {}

        response = self.client.post(
            reverse('marketing:edit_email_for_campaigns', args=(self.contact_email_campaign.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name':'admin name',
            'email':'adminContact@email.com'
        }

        response = self.client.post(
            reverse('marketing:edit_email_for_campaigns', args=(self.contact_email_campaign.id,)), data)
        self.assertEqual(response.status_code, 200)

        data = {
            'name':'admin name',
            'email': self.contact_email_campaign_1.email,
        }

        response = self.client.post(
            reverse('marketing:edit_email_for_campaigns', args=(self.contact_email_campaign.id,)), data)
        self.assertEqual(response.status_code, 200)

    def test_delete_email_for_campaigns(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:delete_email_for_campaigns', args=(self.contact_email_campaign.id,)))
        self.assertEqual(response.status_code, 302)


class TestContactListFileUploadForXlsxAndXls(TestMarketingModel, TestCase):

    def test_contact_list_file_upload_for_xlsx(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_new'))
        self.assertEqual(response.status_code, 200)

        file_headers = ['company name', 'email', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]
        wb = openpyxl.workbook.Workbook()
        dest_filename = 'marketing/test_1.xlsx'
        ws1 = wb.active
        ws1.title = 'test xlsx files'
        ws1.append(file_headers)
        for row in file_content:
            ws1.append(row)
        wb.save(filename=dest_filename)
        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xlsx', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 201)
        os.remove(dest_filename)

        # missing email header
        file_headers = ['company name', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        wb = openpyxl.workbook.Workbook()
        dest_filename = 'marketing/test_1.xlsx'
        ws1 = wb.active
        ws1.title = 'test xlsx files'
        ws1.append(file_headers)
        for row in file_content:
            ws1.append(row)
        wb.save(filename=dest_filename)
        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xlsx', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)
        os.remove(dest_filename)

        # missing all headers
        file_headers = []
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        wb = openpyxl.workbook.Workbook()
        dest_filename = 'marketing/test_1.xlsx'
        ws1 = wb.active
        ws1.title = 'test xlsx files'
        ws1.append(file_headers)
        for row in file_content:
            ws1.append(row)
        wb.save(filename=dest_filename)
        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xlsx', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)
        os.remove(dest_filename)

        # invalid emails
        file_headers = ['company name', 'email', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user1.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user5@email', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]
        wb = openpyxl.workbook.Workbook()
        dest_filename = 'marketing/test_1.xlsx'
        ws1 = wb.active
        ws1.title = 'test xlsx files'
        ws1.append(file_headers)
        for row in file_content:
            ws1.append(row)
        wb.save(filename=dest_filename)
        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xlsx', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)

        os.remove(dest_filename)


    def test_contact_list_file_upload_for_xls(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:contact_list_new'))
        self.assertEqual(response.status_code, 200)

        file_headers = ['company name', 'email', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        dest_filename = 'test_1.xls'
        wb = xlwt.Workbook()
        ws = wb.add_sheet('test xls')
        for count, row in enumerate(file_headers):
            ws.write(0,count,row)

        col_location = 1
        for row in file_content:
            for count_row_content, row_content in enumerate(row):
                ws.write(col_location, count_row_content, row_content)
            col_location = col_location + 1
        wb.save(dest_filename)

        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xls', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 201)
        os.remove(dest_filename)

        # missing email headers
        file_headers = ['company name', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        dest_filename = 'test_1.xls'
        wb = xlwt.Workbook()
        ws = wb.add_sheet('test xls')
        for count, row in enumerate(file_headers):
            ws.write(0,count,row)

        col_location = 1
        for row in file_content:
            for count_row_content, row_content in enumerate(row):
                ws.write(col_location, count_row_content, row_content)
            col_location = col_location + 1
        wb.save(dest_filename)

        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xls', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)
        os.remove(dest_filename)

        # missing headers
        file_headers = []
        file_content = [
            ['company_name_1', 'user1@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user2@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_3', 'user3@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_4', 'user4@email.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        dest_filename = 'test_1.xls'
        wb = xlwt.Workbook()
        ws = wb.add_sheet('test xls')
        for count, row in enumerate(file_headers):
            ws.write(0,count,row)

        col_location = 1
        for row in file_content:
            for count_row_content, row_content in enumerate(row):
                ws.write(col_location, count_row_content, row_content)
            col_location = col_location + 1
        wb.save(dest_filename)

        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xls', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)
        os.remove(dest_filename)

        # invalid emails
        file_headers = ['company name', 'email', 'first name', 'last name', 'city', 'state']
        file_content = [
            ['company_name_1', 'user5@email', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
            ['company_name_2', 'user6.com', 'first_name', 'last_name', 'Hyderabad', 'Telangana'],
        ]

        dest_filename = 'test_1.xls'
        wb = xlwt.Workbook()
        ws = wb.add_sheet('test xls')
        for count, row in enumerate(file_headers):
            ws.write(0,count,row)

        col_location = 1
        for row in file_content:
            for count_row_content, row_content in enumerate(row):
                ws.write(col_location, count_row_content, row_content)
            col_location = col_location + 1
        wb.save(dest_filename)

        with open(dest_filename, 'rb') as fp:
            data = {
                'name': 'sample_test xls', 'contacts_file': fp, 'tags':''
            }
            response = self.client.post(reverse('marketing:contact_list_new'), data)
            self.assertEqual(response.status_code, 200)
        os.remove(dest_filename)


class TestCampaignLinkClick(TestMarketingModel, TestCase):

    def test_campaign_link_click(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_link_click', kwargs={
                'link_id': self.link.id,
                'email_id':self.contact_1.id,
            }))
        self.assertEqual(response.status_code, 302)

        response = self.client.get(
            reverse('marketing:campaign_link_click', kwargs={
                'link_id': self.link.id,
                'email_id':self.contact_1.id,
            }))
        self.assertEqual(response.status_code, 302)

        response = self.client.get(
            reverse('marketing:campaign_link_click', kwargs={
                'link_id': self.link.id,
                'email_id':self.contact_email_campaign.id,
            }))
        self.assertEqual(response.status_code, 302)

    def test_campaign_open(self):
        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_open', kwargs={
                'campaign_log_id': self.campaign.id,
                'email_id':self.contact_1.id,
            }))
        self.assertEqual(response.status_code, 200)

        self.client.login(username='john@example.com', password='password')
        response = self.client.get(
            reverse('marketing:campaign_open', kwargs={
                'campaign_log_id': self.campaign.id,
                'email_id':self.contact_1.id,
            }))
        self.assertEqual(response.status_code, 200)


class TestMarketingModelMethods(TestMarketingModel, TestCase):

    def test_marketing_model_methods(self):
        self.assertEqual(self.tag_marketing.created_by, self.user)
        self.assertEqual(self.tag_marketing_1.created_by, None)
        self.assertEqual(self.email_template.created_by_user, self.user)
        self.assertEqual(self.contact_list.created_by_user, self.user)
        self.assertTrue(self.tag_marketing in self.contact_list.tags_data)
        self.assertEqual(self.contact_list.no_of_contacts, 4)
        self.assertEqual(self.contact_list.no_of_campaigns, 1)
        self.assertEqual(self.contact_list.unsubscribe_contacts, 0)
        self.assertTrue(self.contact_list.created_on_arrow in ['just now', 'seconds ago'])
        self.assertEqual(str(self.contact), self.contact.email)
        self.assertEqual(str(self.failed_contact), self.failed_contact.email)
        self.assertEqual(self.campaign.no_of_clicks, 0)
        self.assertEqual(self.campaign.sent_on_format, self.campaign.created_on.strftime('%b %d, %Y %I:%M %p'))
        self.assertEqual(self.contact_list.created_on_format, self.contact_list.created_on.strftime('%b %d, %Y %I:%M %p'))
        self.assertTrue(self.campaign.sent_on_arrow in ['just now', 'seconds ago'])
        # self.assertEqual(self.campaign_scheduled_later.sent_on_format,
        #     datetime.strptime(self.campaign_scheduled_later.schedule_date_time, '%Y-%m-%d %H:%M'
        #     ).strftime('%b %d, %Y %I:%M %p'))
