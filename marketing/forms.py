import csv
import datetime
import json
import re

import openpyxl
import xlrd
from django import forms

from common.models import User
from marketing.models import (BlockedDomain, BlockedEmail, Campaign, Contact,
                              ContactEmailCampaign, ContactList, EmailTemplate, Tag)
from haystack.forms import SearchForm

email_regex = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"


def csv_doc_validate(document):
    temp_row = []
    invalid_row = []
    # this stores all the failed csv contacts
    failed_contacts_csv = []
    reader = csv.reader((document.read().decode("iso-8859-1")).splitlines())
    # csv_headers = ["first name", "last name", "email"]
    csv_headers = ["first name", "email"]
    # required_headers = ["first name", "last name", "email"]
    required_headers = ["first name", "email"]
    for y_index, row in enumerate(reader):
        each = {}
        invalid_each = {}
        if y_index == 0:
            csv_headers = [header_name.lower()
                           for header_name in row if header_name]
            missing_headers = set(required_headers) - \
                set([r.lower() for r in row])
            if missing_headers:
                missing_headers_str = ', '.join(missing_headers)
                message = 'Missing headers: %s' % (missing_headers_str)
                return {"error": True, "message": message}
            continue
        elif not ''.join(str(x) for x in row):
            continue
        else:
            for x_index, cell_value in enumerate(row):
                try:
                    csv_headers[x_index]
                except IndexError:
                    continue
                if csv_headers[x_index] in required_headers:
                    if not cell_value:
                        # message = 'Missing required value %s for row %s' % (
                        #     csv_headers[x_index], y_index + 1)
                        # return {"error": True, "message": message}
                        invalid_each[csv_headers[x_index]] = cell_value
                    else:
                        if csv_headers[x_index] == "email":
                            if re.match(email_regex, cell_value) is None:
                                invalid_each[csv_headers[x_index]] = cell_value
                each[csv_headers[x_index]] = cell_value
        if invalid_each:
            invalid_row.append(each)
            failed_contacts_csv.append(list(each.values()))
        else:
            temp_row.append(each)
    return {"error": False, "validated_rows": temp_row, "invalid_rows": invalid_row, "headers": csv_headers,
            "failed_contacts_csv": failed_contacts_csv}


# def get_validated_rows(wb, sheet_name, validated_rows, invalid_rows):
#     # headers = ["first name", "last name", "email"]
#     # required_headers = ["first name", "last name", "email"]
#     headers = ["first name", "email"]
#     required_headers = ["first name", "email"]
#     work_sheet = wb.get_sheet_by_name(name=sheet_name)
#     for y_index, row in enumerate(work_sheet.iter_rows()):
#         if y_index == 0:
#             missing_headers = set(required_headers) - \
#                 set([str(cell.value).lower() for cell in row])
#             if missing_headers:
#                 missing_headers_str = ', '.join(missing_headers)
#                 message = "Missing headers: %s %s" % (
#                     missing_headers_str, sheet_name)
#                 return {"error": True, "message": message}
#             continue
#         elif not ''.join(str(cell.value) for cell in row):
#             continue
#         else:
#             temp_row = []
#             invalid_row = []
#             for x_index, cell in enumerate(row):
#                 try:
#                     headers[x_index]
#                 except IndexError:
#                     continue
#                 if headers[x_index] in required_headers:
#                     if not cell.value:
#                         # message = 'Missing required \
#                         #             value %s for row %s in sheet %s'\
#                         #     % (headers[x_index], y_index + 1, sheet_name)
#                         # return {"error": True, "message": message}
#                         invalid_row.append(headers[x_index])
#                 temp_row.append(cell.value)
#             if len(invalid_row) > 0:
#                 invalid_rows.append(temp_row)
#             else:
#                 if len(temp_row) >= len(required_headers):
#                     validated_rows.append(temp_row)
#     return validated_rows, invalid_rows

def get_validated_rows(wb, sheet_name, validated_rows, invalid_rows):
    wb_sheet = wb[sheet_name]
    sheet_headers = [cell.value for cell in wb_sheet[1]]
    required_headers = ["first name", "email"]

    # this condition will check if the both fields in required_headers exist
    if not len(set(sheet_headers).intersection(required_headers)) == 2:
        missing_headers = set(required_headers) - set(sheet_headers)
        missing_headers_str = ', '.join(missing_headers)
        message = 'Missing headers: {}'.format(missing_headers_str)
        return {"error": True, "message": message}, message
    else:
        data = []
        for row in wb_sheet.rows:
            d = {}
            for key, cell_value in zip(sheet_headers, row):
                d[key] = cell_value.value
            data.append(d)
        # remove the first element, it contains no values
        data.pop(0)
        if len(data):
            validated_rows = []
            invalid_rows = []
            for row in data:
                if row.get('email', None):
                    if re.match(email_regex, row.get('email', None)) is None:
                        invalid_rows.append(row)
                    else:
                        validated_rows.append(row)
            return validated_rows, invalid_rows


def xlsx_doc_validate(document):
    wb = openpyxl.load_workbook(document)
    sheets = wb.sheetnames
    validated_rows = []
    invalid_rows = []
    for sheet_name in sheets:
        valid_data, invalid_data = get_validated_rows(
            wb, sheet_name, validated_rows, invalid_rows)
        if type(valid_data) == dict:
            return valid_data
        validated_rows = validated_rows + valid_data
        invalid_rows = invalid_rows + invalid_data
    return {"error": False, "validated_rows": validated_rows, "invalid_rows": invalid_rows}


def get_validated_rows_xls(wb, sheet_name, validated_rows, invalid_rows):
    wb_sheet = wb.sheet_by_name(sheet_name)
    sheet_headers = [cell.value for cell in wb_sheet.row(0)]
    required_headers = ["first name", "email"]

    # this condition will check if the both fields in required_headers exist
    if not len(set(sheet_headers).intersection(required_headers)) == 2:
        missing_headers = set(required_headers) - set(sheet_headers)
        missing_headers_str = ', '.join(missing_headers)
        message = 'Missing headers: {}'.format(missing_headers_str)
        return {"error": True, "message": message}, message
    else:
        no_of_rows = wb_sheet.nrows
        data = []
        for nrow in range(no_of_rows):
            d = {}
            for key, cell_value in zip(sheet_headers, [cell.value for cell in wb_sheet.row(nrow)]):
                d[key] = cell_value
            data.append(d)
        # remove the first element, it contains no values
        data.pop(0)
        if len(data):
            validated_rows = []
            invalid_rows = []
            for row in data:
                if row.get('email', None):
                    if re.match(email_regex, row.get('email', None)) is None:
                        invalid_rows.append(row)
                    else:
                        validated_rows.append(row)
            return validated_rows, invalid_rows


def xls_doc_validate(document):
    wb = xlrd.open_workbook(file_contents=document.open().read())
    sheets = wb.sheet_names()
    validated_rows = []
    invalid_rows = []
    for sheet_name in sheets:
        valid_data, invalid_data = get_validated_rows_xls(
            wb, sheet_name, validated_rows, invalid_rows)
        # if the data is valid data we'll get a list or else a dict
        if type(valid_data) == dict:
            return valid_data
        validated_rows = validated_rows + valid_data
        invalid_rows = invalid_rows + invalid_data
        return {"error": False, "validated_rows": validated_rows, "invalid_rows": invalid_rows}


def import_document_validator(document):
    try:
        # dialect = csv.Sniffer().sniff(document.read(1024).decode("ascii"))
        document.seek(0, 0)  # csv file
        return csv_doc_validate(document)
    except Exception as e:
        print(e)
        print('csv')
        try:
            return xlsx_doc_validate(document)  # xlsx file
        except Exception as e:
            print(e)
            print('xlsx')
            try:
                return xls_doc_validate(document)  # xls file
            except Exception as e:
                print(e)
                print('xls')
                return {"error": True, "message": "Not a valid CSV/XLS, XLSX file"}


class ContactListForm(forms.ModelForm):
    tags = forms.CharField(max_length=5000, required=False)
    contacts_file = forms.FileField(required=False)

    class Meta:
        model = ContactList
        fields = ["name"]

    def __init__(self, *args, **kwargs):
        super(ContactListForm, self).__init__(*args, **kwargs)
        self.fields['contacts_file'].widget.attrs.update({
            "accept": ".csv,.xls,.xlsx,.xlsm,.xlsb,.xml",
        })
        if self.instance.id is None:
            self.fields['contacts_file'].required = True
        else:
            self.fields['contacts_file'].required = False
        if self.data.get('contacts_file'):
            self.fields['contacts_file'].widget.attrs.update({
                "accept": ".csv,.xls,.xlsx,.xlsm,.xlsb,.xml",
            })

    def clean_name(self):
        name = self.cleaned_data.get('name')
        if ContactList.objects.filter(name__iexact=name).exclude(id=self.instance.id).exists():
            raise forms.ValidationError(
                'Contact List with this Name already exists.')
        return name

    def clean_contacts_file(self):
        document = self.cleaned_data.get("contacts_file")
        if document:
            data = import_document_validator(document)
            if data.get("error"):
                raise forms.ValidationError(data.get("message"))
            else:
                self.validated_rows = data.get("validated_rows", [])
                self.invalid_rows = data.get("invalid_rows", [])
                if len(self.validated_rows) == 0:
                    raise forms.ValidationError(
                        "All the contacts in the file are invalid.")
                # self.headers = data.get("headers", [])
                # self.failed_contacts_csv = data.get("failed_contacts_csv", [])
                # failed_csv_data = []
                # failed_csv_data = (self.failed_contacts_csv)
                # failed_csv_data.insert(0, self.headers)
                # if self.invalid_rows:
                #     self.add_error(None, json.dumps(failed_csv_data))
                #     raise forms.ValidationError('Uploaded file is not valid.')
        return document

    # def clean_visible_to(self):
    #     visible_to_data = json.loads(self.data['visible_to'])
    #     if len(visible_to_data) > 0:
    #         instance_visible_to = []
    #         if self.instance and self.instance.id is not None:
    #             instance_visible_to = list(
    #                 set(self.instance.visible_to.all().values_list(
    #                     'email', flat=True)))
    #         for each in visible_to_data:
    #             visible_to = User.objects.filter(email=each)
    #             if instance_visible_to:
    #                 visible_to = visible_to.exclude(
    #                     email__in=instance_visible_to)
    #             if visible_to:
    #                 raise forms.ValidationError(
    #                     str(each) + ' User aleady existed')
    #         return self.data['visible_to']
    #     raise forms.ValidationError("Select any of the users")


class ContactForm(forms.ModelForm):
    contact_list = forms.CharField(max_length=5000)

    def __init__(self, *args, **kwargs):
        request_user = kwargs.pop('request_user', None)
        self.obj_instance = kwargs.get('instance', None)
        super(ContactForm, self).__init__(*args, **kwargs)
        for field in self.fields.values():
            field.widget.attrs = {"class": "form-control"}

        self.fields['name'].required = True
        self.fields['email'].required = True
        self.fields['last_name'].required = False
        self.fields['city'].required = False
        self.fields['state'].required = False
        self.fields['company_name'].required = False
        self.fields['contact_list'].required = False

    class Meta:
        model = Contact
        fields = ["name", "email", "contact_number",
                  "last_name", "city", "state", "company_name"]

    def clean_email(self):
        email = self.cleaned_data.get('email')
        if Contact.objects.filter(email=email).exclude(id=self.instance.id).exists():
            raise forms.ValidationError(
                'Contact with this Email already exists')
        return email

    # def clean_contact_list(self):
    #     contact_list = self.cleaned_data.get("contact_list")
    #     if not contact_list or contact_list == '[]' or\
    #             json.loads(contact_list) == []:
    #         raise forms.ValidationError(
    #             "Please choose any of the Contact List")
    #     else:
    #         for each in json.loads(contact_list):
    #             if not ContactList.objects.filter(id=each).first():
    #                 raise forms.ValidationError(
    #                     "Please choose a valid Contact List")

    #     return contact_list


# class ContactsCSVUploadForm(forms.Form):
#     contacts_file = forms.FileField()
#     contact_list = forms.CharField(max_length=5000)

#     def __init__(self, *args, **kwargs):
#         super(ContactsCSVUploadForm, self).__init__(*args, **kwargs)
#         self.fields['contacts_file'].widget.attrs.update({
#             "accept": ".csv,.xls,.xlsx,.xlsm,.xlsb,.xml",
#         })
#         self.fields['contacts_file'].required = True

#     def clean_contacts_file(self):
#         document = self.cleaned_data.get("contacts_file")
#         data = import_document_validator(document)
#         if data.get("error"):
#             raise forms.ValidationError(data.get("message"))
#         else:
#             self.validated_rows = data.get("validated_rows")
#         return document

#     def clean_contact_list(self):
#         contact_list = self.cleaned_data.get("contact_list")
#         if not contact_list or contact_list == '[]' or\
#                 json.loads(contact_list) == []:
#             raise forms.ValidationError(
#                 "Please choose any of the Contact List")
#         else:
#             for each in json.loads(contact_list):
#                 if not ContactList.objects.filter(id=each).first():
#                     raise forms.ValidationError(
#                         "Please choose a valid Contact List")

#         return contact_list


class EmailTemplateForm(forms.ModelForm):

    class Meta:
        model = EmailTemplate
        fields = ['title', 'subject', 'html']


class SendCampaignForm(forms.ModelForm):
    schedule_later = forms.BooleanField(required=False)
    reply_to_crm = forms.BooleanField(required=False)
    timezone = forms.CharField(max_length=500, required=False)
    schedule_date_time = forms.CharField(max_length=100, required=False)
    reply_to_email = forms.EmailField(max_length=100, required=False)
    from_email = forms.EmailField(max_length=100, required=True)
    from_name = forms.CharField(max_length=100, required=True)
    contact_list = forms.CharField(max_length=5000, required=True)
    tags = forms.CharField(max_length=5000, required=False)

    class Meta:
        model = Campaign
        fields = ['title', 'subject', 'html', 'email_template', 'attachment', ]

    def __init__(self, *args, **kwargs):
        self.contacts_list = kwargs.pop('contacts_list')
        super(SendCampaignForm, self).__init__(*args, **kwargs)
        if self.data.get('schedule_later') and self.data['schedule_later'] == 'true':
            self.fields['timezone'].required = True
            self.fields['schedule_date_time'].required = True
        if not self.data.get('reply_to_crm'):
            self.fields['reply_to_email'].required = True

    def clean_schedule_date_time(self):
        schedule_date_time = self.cleaned_data.get('schedule_date_time')
        if self.cleaned_data.get('schedule_later') == True:
            schedule_date_time = datetime.datetime.strptime(
                schedule_date_time, '%Y-%m-%d %H:%M')
            if schedule_date_time < datetime.datetime.now():
                raise forms.ValidationError(
                    'Schedule Date Time should be greater than current time')
        return schedule_date_time

    def clean_contact_list(self):
        contact_list = self.contacts_list
        if not contact_list or contact_list == '[]' or \
                contact_list == []:
            raise forms.ValidationError(
                "Please choose any of the Contact List")
        else:
            for each in contact_list:
                contacts_list_obj = ContactList.objects.filter(id=each).first()
                if not contacts_list_obj:
                    raise forms.ValidationError(
                        "Please choose a valid Contact List")

                if contacts_list_obj.contacts.count() < 1:
                    raise forms.ValidationError(
                        'The contact list "{}" does not have any contacts in it .'.format(contacts_list_obj.name))

        return contact_list

    def clean_html(self):
        html = self.cleaned_data.get('html')
        count = 0
        for i in html:
            if i == "{":
                count += 1
            elif i == "}":
                count -= 1
            if count < 0:
                raise forms.ValidationError(
                    'Brackets do not match, Enter valid tags.')
        if count != 0:
            raise forms.ValidationError(
                'Brackets do not match, Enter valid tags.')
        return html

    def clean_title(self):
        title = self.cleaned_data.get('title')
        if Campaign.objects.filter(title__iexact=title).exclude(id=self.instance.id).exists():
            raise forms.ValidationError(
                'Campaign with this title already exists.')
        return title


class EmailCampaignForm(forms.ModelForm):

    def __init__(self, *args, **kwargs):
        self.obj_instance = kwargs.get('instance', None)
        super(EmailCampaignForm, self).__init__(*args, **kwargs)
        for field in self.fields.values():
            field.widget.attrs = {"class": "form-control"}

        self.fields['name'].required = True
        self.fields['email'].required = True
        self.fields['last_name'].required = False

    class Meta:
        model = ContactEmailCampaign
        fields = ["name", "email", "last_name", ]

    def clean_email(self):
        email = self.cleaned_data.get('email')
        if ContactEmailCampaign.objects.filter(email=email).exclude(id=self.instance.id).exists():
            raise forms.ValidationError(
                'Contact with this Email already exists.')
        return email


class BlockedDomainsForm(forms.ModelForm):

    class Meta:
        model = BlockedDomain
        fields = ['domain', ]

    def clean_domain(self):
        domain = self.cleaned_data.get('domain')
        domain_regex = '^([a-z0-9]+(-[a-z0-9]+)*\.)+[a-z]{2,}$'
        if re.match(domain_regex, domain) is None:
            raise forms.ValidationError('Enter a valid domain.')
        if BlockedDomain.objects.filter(domain__iexact=domain).exclude(id=self.instance.id).exists():
            raise forms.ValidationError('Domain with this name already exists.')
        return domain

class BlockedEmailForm(forms.ModelForm):

    class Meta:
        model = BlockedEmail
        fields = ['email', ]

    def clean_email(self):
        email = self.cleaned_data.get('email')
        if BlockedEmail.objects.filter(email__iexact=email).exclude(id=self.instance.id).exists():
            raise forms.ValidationError('Email already exists.')
        return email


class MarketingContactEmailSearchForm(SearchForm):
    email_domain = forms.CharField()

    def search(self):
        # First, store the SearchQuerySet received from other processing.
        sqs = super(MarketingContactEmailSearchForm, self).search()

        if not self.is_valid():
            return self.no_query_found()

        # Check to see if a start_date was chosen.
        if self.cleaned_data['email_domain']:
            # import pdb; pdb.set_trace()
            sqs = sqs.filter(email__icontains=self.cleaned_data.get('email_domain'))

        return sqs

    def clean_email_domain(self):
        email_domain = self.cleaned_data.get('email_domain')
        domain_regex = '^([a-z0-9]+(-[a-z0-9]+)*\.)+[a-z]{2,}$'
        if re.match(domain_regex, email_domain) is None:
            raise forms.ValidationError('Enter a valid domain.')
        return email_domain